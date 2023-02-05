# -*- coding: utf-8 -*-
import os
import textwrap
import random
import time
import base64

import cv2
import hashlib
import requests
import openpyxl as xl
from openpyxl.styles import Alignment
from prettytable import PrettyTable
import argparse


def get_user_agent():
    """
    获取随机UA头
    """
    user_agent_list = [
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1" \
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5", \
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5", \
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
    ]
    user_agent = random.choice(user_agent_list)
    return user_agent


def get_cookies():
    cookie_headers = {'User-Agent': get_user_agent()}
    for i in range(5):
        try:
            cookie = requests.utils.dict_from_cookiejar(
                requests.get('https://beian.miit.gov.cn/',
                             headers=cookie_headers).cookies)['__jsluid_s']
            if method == 'single':
                print('[√] 获取 cookie 成功')
            return cookie
        except:
            time.sleep(0.5)
    raise ValueError('[×] 获取Cookie失败，请稍后再试！')


def get_token():
    timeStamp = round(time.time() * 1000)
    authSecret = 'testtest' + str(timeStamp)
    authKey = hashlib.md5(authSecret.encode(encoding='UTF-8')).hexdigest()
    auth_data = {'authKey': authKey, 'timeStamp': timeStamp}
    url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/auth'
    for i in range(5):
        try:
            t_response = requests.post(url=url, data=auth_data, headers=base_header).json()
            token = t_response['params']['bussiness']
            if method == 'single':
                print('[√] 获取 token 成功')
            return token
        except:
            time.sleep(0.5)
    raise ValueError('[×] 获取Token失败，请稍后再试！')


def get_check_pic(token):
    url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/image/getCheckImage'
    base_header['Accept'] = 'application/json, text/plain, */*'
    base_header.update({'Content-Length': '0', 'token': token})
    for i in range(5):
        try:
            p_request = requests.post(url=url, data='', headers=base_header).json()
            p_uuid = p_request['params']['uuid']
            big_image = p_request['params']['bigImage']
            small_image = p_request['params']['smallImage']
            # 解码图片，写入并计算图片缺口位置
            with open('bigImage.jpg', 'wb') as f:
                f.write(base64.b64decode(big_image))
            with open('smallImage.jpg', 'wb') as f:
                f.write(base64.b64decode(small_image))
            background_image = cv2.imread('bigImage.jpg', cv2.COLOR_GRAY2RGB)
            fill_image = cv2.imread('smallImage.jpg', cv2.COLOR_GRAY2RGB)
            position_match = cv2.matchTemplate(background_image, fill_image, cv2.TM_CCOEFF_NORMED)
            max_loc = cv2.minMaxLoc(position_match)[3][0]
            mouse_length = max_loc + 1
            os.remove('bigImage.jpg')
            os.remove('smallImage.jpg')
            check_data = {'key': p_uuid, 'value': mouse_length}
            if method == 'single':
                print('[√] 计算验证码滑块缺口位置成功')
            return check_data
        except:
            time.sleep(0.5)
    raise ValueError('[×] 计算验证码滑块缺口位置错误，请稍后再试！')


def get_sign(check_data, token):
    check_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/image/checkImage'
    base_header.update({'Content-Length': '60', 'token': token, 'Content-Type': 'application/json'})
    for i in range(5):
        try:
            pic_sign = requests.post(check_url, json=check_data, headers=base_header).json()
            sign = pic_sign['params']
            if method == 'single':
                print('[√] 自动破解滑块验证码成功')
            return sign
        except:
            time.sleep(0.5)
    raise ValueError('[×] 图片验证码破解失败，请稍后再试！')


def get_icp_info(info_data, p_uuid, token, sign):
    """
    获取备案信息
    """
    domain_list = []
    info_url = 'https://hlwicpfwc.miit.gov.cn/icpproject_query/api/icpAbbreviateInfo/queryByCondition'
    base_header.update({'Content-Length': '78', 'uuid': p_uuid, 'token': token, 'sign': sign})
    # 请求获取备案信息
    try:
        icp_info = requests.post(url=info_url, json=info_data, headers=base_header).json()
        if icp_info['code'] == 429:
            print("查询失败，{}".format(icp_info["msg"]))
            return domain_list
    except:
        for i in range(3):
            time.sleep(0.5)
            icp_info = requests.post(url=info_url, json=info_data, headers=base_header).json()
    # 处理备案信息
    info = info_data['unitName']
    domain_total = icp_info['params']['total']
    page_total = icp_info['params']['lastPage']
    end_row = icp_info['params']['endRow']
    if method == 'single':
        print("查询对象：{} 共有 {} 个已备案域名".format(info, domain_total))
    else:
        print("查询对象共发现 {} 个已备案域名".format(domain_total))
    for i in range(0, page_total):
        print(f"正在查询第{i + 1}页...")
        for k in range(0, end_row + 1):
            info_base = icp_info['params']['list'][k]
            domain_name = info_base['domain']
            domain_type = info_base['natureName']
            domain_licence = info_base['mainLicence']
            website_licence = info_base['serviceLicence']
            domain_status = info_base['limitAccess']
            domain_approve_date = info_base['updateRecordTime']
            domain_owner = info_base['unitName']
            try:
                domain_content_approved = info_base['contentTypeName']
                if domain_content_approved == "":
                    domain_content_approved = "无"
            except KeyError:
                domain_content_approved = "无"
            row_data = domain_owner, domain_name, domain_licence, website_licence, domain_type, domain_content_approved, domain_status, domain_approve_date
            domain_list.append(row_data)
        info_data = {'pageNum': i + 2, 'pageSize': '40', 'unitName': info}
        if icp_info['params']['isLastPage'] is True:
            break
        else:
            icp_info = requests.post(info_url, json=info_data, headers=base_header).json()
            end_row = icp_info['params']['endRow']
            time.sleep(1)
    return domain_list


def query(target):
    # 对传入参数进行过滤
    target = target.strip().replace("https://www.", "").replace("http://www.", "").replace("http://", "")
    query_data = {'pageNum': '1', 'pageSize': '40', 'unitName': target}
    # 获取Cookie值
    cookie = get_cookies()
    global base_header
    base_header = {
        'User-Agent': get_user_agent(),
        'Origin': 'https://beian.miit.gov.cn',
        'Referer': 'https://beian.miit.gov.cn/',
        'Cookie': f'__jsluid_s={cookie}'
    }
    # 获取token
    token = get_token()
    # 滑块验证码识别
    check_data = get_check_pic(token)
    sign = get_sign(check_data, token)
    # 请求备案数据
    p_uuid = check_data['key']
    domain_list = get_icp_info(query_data, p_uuid, token, sign)
    return domain_list


def save_to_excel(domain_list, file_name):
    # 计算需要写入表格的总行数，如果是空列表，即代表该域名没有备案信息，也有可能是获取信息失败了
    total_row = len(domain_list)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return
    print(f"查询结果如下:\n{domain_list}")
    # 判断文件保存路径是否存在，如果不存在则创建
    if not os.path.exists("out"):
        os.mkdir("out")
    else:
        # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
        file_path = "out/" + file_name + ".xlsx"
        if os.path.exists(file_path):
            wb = xl.load_workbook(file_path)
            ws = wb[file_name]
            max_row = ws.max_row
            start = max_row + 1
            total_row = total_row + start
            after_title = 0
        else:
            wb = xl.Workbook()
            ws = wb.active
            ws.title = file_name
            title_list = ['域名主办方', '域名', '备案许可证号', '网站备案号', '域名类型', '网站前置审批项', '是否限制接入', '审核通过日期']
            for i in range(0, 8):
                ws.cell(1, i + 1).value = title_list[i]
            col_width = {'A': 45, 'B': 40, 'C': 22, 'D': 24, 'E': 9, 'F': 15, 'G': 13, 'H': 21}
            for k, v in col_width.items():
                ws.column_dimensions[k].width = v
            ws.freeze_panes = 'A2'
            start = 0
            after_title = 2
        # 写入查询数据
        for j in range(start, total_row + 1):
            for k in range(0, 8):
                try:
                    ws.cell(j + after_title, k + 1).value = domain_list[j - start][k]
                except:
                    continue
        # 垂直居中
        for row in range(ws.max_row):
            for col in range(ws.max_column):
                ws.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center')
        try:
            wb.save(file_path)
        except PermissionError as e:
            print("[×] 目标文件已打开，无法写入数据，请关闭文件后重新执行！")
            return -1
        print("查询结果保存在：{}".format(file_path))


def print_table(domain_list):
    """
    以表格的形式对数据进行整理输出
    """
    if len(domain_list) == 0:
        return
    table = PrettyTable(['域名主办方', '网站备案号', '域名', '域名类型', '审批项', '限制接入', '备案日期'])
    for list in domain_list:
        table.add_row([list[0], list[3], list[1], list[4], list[5], list[6], list[7].split(" ")[0]])
    print(table)


def cmdline():
    banner = """\
       __     ______     ______   ______     ______     __        
      /\ \   /\  ___\   /\  == \ /\  ___\   /\  __ \   /\ \     
      \ \ \  \ \ \____  \ \  _-/ \ \ \____  \ \ \/\ \  \ \ \____  
       \ \_\  \ \_____\  \ \_\    \ \_____\  \ \_____\  \ \_____\ 
        \/_/   \/_____/   \/_/     \/_____/   \/_____/   \/_____/  by OssianSong"""
    parser = argparse.ArgumentParser(
        prog="python ICPCol.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent(banner),
        exit_on_error=False,
        add_help=False)
    group = parser.add_mutually_exclusive_group()
    group.add_argument("-t", metavar="TARGET", type=str, help="输入公司名/备案号/域名/IP查询对应备案信息")
    group.add_argument("-f", metavar="FILE", type=str, help="从指定txt文件中读取数据进行批量查询")
    parser.add_argument("-o", metavar="FILENAME", nargs='?', default=argparse.SUPPRESS,
                        help="指定文件名将数据保存到Excel中,参数为空时使用默认文件名")
    parser.add_argument('-h', '--help', action='help', default=argparse.SUPPRESS, help='查看程序使用帮助')
    return parser


if __name__ == '__main__':
    # 选择输出模式
    global method
    # 获取输入参数
    parser = cmdline()
    args = parser.parse_args()
    domain_list = []
    # 没有参数打印帮助信息
    if not any(vars(args).values()):
        parser.parse_args(["-h"])
    # 单个查询
    if args.t:
        method = 'single'
        domain_list = query(args.t)
        # 对输出结果进行处理
        if "o" not in args:
            print_table(domain_list)
        elif args.o:
            save_to_excel(domain_list, args.o)
        else:
            file_name = domain_list[0][0]
            save_to_excel(domain_list, file_name)
    # 多个查询
    elif args.f:
        try:
            method = 'multiple'
            count = 1
            with open(args.f, 'r', encoding='utf-8') as f:
                for target in f.readlines():
                    print("[{}] 查询对象：{}".format(count, target.strip()))
                    domain_list = query(target.strip())
                    count += 1
                    # 对输出结果进行处理
                    if "o" not in args:
                        print_table(domain_list)
                    elif args.o:
                        save_to_excel(domain_list, args.o)
                    else:
                        file_name = time.strftime('%Y%m%d%H%M%S', time.localtime())
                        save_to_excel(domain_list, file_name)
        except Exception as e:
            print(e)
            print("[×] 打开文件不存在，请确认文件位置！")
    else:
        raise ValueError("[×] 请先指定查询目标！")