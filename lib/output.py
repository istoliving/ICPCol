import os
import re

import openpyxl as xl
from openpyxl.styles import Alignment
from prettytable import PrettyTable


def is_valid_ip(ip):
    """
    验证字符串是否为IP地址
    """
    pattern = re.compile(r"^(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$")
    if pattern.match(ip):
        return True
    return False


def save_to_excel(domain_list, file_name):
    # 计算需要写入表格的总行数，如果是空列表，即代表该域名没有备案信息，也有可能是获取信息失败了
    total_row = len(domain_list)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return
    print(domain_list)
    # 判断文件保存路径是否存在，如果不存在则创建
    if not os.path.exists("out"):
        os.mkdir("out")
    else:
        # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
        file_path = "out/" + file_name + ".xlsx"
        if os.path.exists(file_path):
            wb = xl.load_workbook(file_path)
            ws = wb[file_name]
            max_row = ws.max_row
            start = max_row + 1
            total_row = total_row + start
            after_title = 0
        else:
            wb = xl.Workbook()
            ws = wb.active
            ws.title = file_name
            title_list = ['域名主办方', '域名', '备案许可证号', '网站备案号', '域名类型', '网站前置审批项', '是否限制接入', '审核通过日期']
            for i in range(0, 8):
                ws.cell(1, i + 1).value = title_list[i]
            col_width = {'A': 45, 'B': 40, 'C': 22, 'D': 24, 'E': 9, 'F': 15, 'G': 13, 'H': 21}
            for k, v in col_width.items():
                ws.column_dimensions[k].width = v
            ws.freeze_panes = 'A2'
            start = 0
            after_title = 2
        # 写入查询数据
        for j in range(start, total_row + 1):
            for k in range(0, 8):
                try:
                    ws.cell(j + after_title, k + 1).value = domain_list[j - start][k]
                except:
                    continue
        # 垂直居中
        for row in range(ws.max_row):
            for col in range(ws.max_column):
                ws.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center')
        try:
            wb.save(file_path)
        except PermissionError as e:
            print("[×] 目标文件已打开，无法写入数据，请关闭文件后重新执行！")
            return -1
        print("查询结果保存在：{}".format(file_path))


def print_table(domain_list):
    """
    以表格的形式对数据进行整理输出
    """
    if len(domain_list) == 0:
        return
    table = PrettyTable(['域名主办方', '网站备案号', '域名', '域名类型', '审批项', '限制接入', '备案日期'])
    for data in domain_list:
        table.add_row([data[0], data[3], data[1], data[4], data[5], data[6], data[7].split(" ")[0]])
    print(table)


def save_to_txt(domain_list, file_name):
    """
    将查询结果保存到txt文件中
    :param domain_list:
    :param file_name:
    :return:
    """
    if len(domain_list) == 0:
        print("查询结果为空")
        return
    print(domain_list)
    ext = file_name.split(".")[-1]
    if ext == "txt":
        file_name = file_name.replace(".txt", "")
    # 判断文件保存路径是否存在，如果不存在则创建
    if not os.path.exists("out"):
        os.mkdir("out")
    file_path = "out/{}.txt".format(file_name)
    with open(file_path, 'a', encoding='utf-8') as f:
        for data in domain_list:
            if is_valid_ip(data[1]):
                f.write('ip="{}"\n'.format(data[1]))
            else:
                f.write('domain="{}"\n'.format(data[1]))
    print("查询语法保存在：{}".format(file_path))
